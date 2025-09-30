# -*- coding: utf-8 -*-
import os
import io
import csv
import time
import json
import requests
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo
from html import escape as h

from app_config import BOT_TOKEN, TZ, VADIM_CHAT_ID, ASSISTANT_CHAT_IDS
from db import (
    get_conn,
    tasks_sent_between,
    find_open_tasks_for_user,
    list_unique_assignees,
    get_reassignments_between,
    get_deadline_changes_between,
    get_deadline_changes_for_task,
    enqueue_outbox,
    pop_due_outbox,
    mark_outbox_sent,
    get_closed_tasks_between,
    count_open_like,
    count_closed_between,
    get_overdue_open_tasks,
)

# --- опционально: openpyxl для старого XLSX-отчёта (пусть остаётся для обратной совместимости)
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
except Exception:
    openpyxl = None

# --- PDF: reportlab (Unicode шрифты)
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import sys, os

    _FONT_READY = False
    def _ensure_pdf_font():
        """
        Регистрируем Unicode TTF.
        Приоритет: ./fonts/NotoSans -> /usr/share/.../NotoSans -> ./fonts/DejaVuSans -> системная DejaVu -> ./fonts/Roboto -> системная Roboto.
        Логируем причины пропуска.
        """
        global _FONT_READY
        if _FONT_READY:
            return _FONT_READY  # (regular_name, bold_name)

        base = os.path.dirname(__file__)
        paths = [
            ("NotoSans",      os.path.join(base, "fonts", "NotoSans-Regular.ttf")),
            ("NotoSans-Bold", os.path.join(base, "fonts", "NotoSans-Bold.ttf")),
            ("NotoSans",      "/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf"),
            ("NotoSans-Bold", "/usr/share/fonts/truetype/noto/NotoSans-Bold.ttf"),

            ("DejaVuSans",      os.path.join(base, "fonts", "DejaVuSans.ttf")),
            ("DejaVuSans-Bold", os.path.join(base, "fonts", "DejaVuSans-Bold.ttf")),
            ("DejaVuSans",      "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
            ("DejaVuSans-Bold", "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),

            ("Roboto",      os.path.join(base, "fonts", "Roboto-Regular.ttf")),
            ("Roboto-Bold", os.path.join(base, "fonts", "Roboto-Bold.ttf")),
            ("Roboto",      "/usr/share/fonts/truetype/roboto/Roboto-Regular.ttf"),
            ("Roboto-Bold", "/usr/share/fonts/truetype/roboto/Roboto-Bold.ttf"),
        ]

        found = set()
        last_err = None
        for name, path in paths:
            if name in found:
                continue
            try:
                if os.path.exists(path) and os.path.getsize(path) > 1024:
                    pdfmetrics.registerFont(TTFont(name, path))
                    print(f"[fonts] loaded {name}: {path}", file=sys.stderr)
                    found.add(name)
                else:
                    if os.path.exists(path):
                        print(f"[fonts] skip {name}: too small ({os.path.getsize(path)} B) at {path}", file=sys.stderr)
                    else:
                        print(f"[fonts] not found {name}: {path}", file=sys.stderr)
            except Exception as e:
                last_err = e
                print(f"[fonts] register fail {name} at {path}: {e}", file=sys.stderr)

        for reg, bld in (("NotoSans","NotoSans-Bold"),
                         ("DejaVuSans","DejaVuSans-Bold"),
                         ("Roboto","Roboto-Bold")):
            if reg in found:
                if bld not in found:
                    bld = reg
                _FONT_READY = (reg, bld)
                return _FONT_READY

        if last_err:
            print(f"[fonts] no usable TTF, last error: {last_err}", file=sys.stderr)
        return None, None

except Exception:
    canvas = None
    def _ensure_pdf_font():
        return None, None


TG_API = f"https://api.telegram.org/bot{BOT_TOKEN}"
TZINFO = ZoneInfo(TZ)

WORK_START = 9   # 09:00
WORK_END   = 18  # 18:00
WEEKDAYS = {0,1,2,3,4}  # пн-пт

ASSIGNEE_DELAY_S   = 0.20   # пауза на каждого исполнителя
ASSISTANTS_BEGIN_PAUSE_S = 2.0
ASSISTANT_DELAY_S  = 1.25   # пауза между ассистентами
ASSIGNEE_BATCH_SIZE = 25    # после каждых 25 — микро-пауза
ASSIGNEE_BATCH_PAUSE_S = 1.0

REPORT_DIR = "/tmp"  # куда класть xlsx/csv/pdf

def is_work_time(dt):
    local = dt.astimezone(TZINFO)
    # Разрешаем ровно в 18:00 отправку дайджеста (минуты == 0)
    return (local.weekday() in WEEKDAYS) and (
        (WORK_START <= local.hour < WORK_END) or (local.hour == WORK_END and local.minute == 0)
    )

def next_work_morning(dt):
    local = dt.astimezone(TZINFO)
    # Ближайшее 09:00 рабочего дня
    d = local
    if local.hour >= WORK_END or local.hour < WORK_START:
        d = local.replace(hour=WORK_START, minute=0, second=0, microsecond=0)
        if local.hour >= WORK_END:
            d += timedelta(days=1)
    while d.weekday() not in WEEKDAYS:
        d = (d + timedelta(days=1)).replace(hour=WORK_START, minute=0, second=0, microsecond=0)
    return d.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

def send_or_enqueue(chat_id, text, markup=None, *, base_delay_s: float = 0.0):
    now = datetime.now(TZINFO)
    if is_work_time(now):
        send(chat_id, text, markup, base_delay_s=base_delay_s)
    else:
        enqueue_outbox(chat_id, text, markup, next_work_morning(now))

def format_postponed(t):
    try:
        postponed = int((t["postponed"] or 0))
    except Exception:
        postponed = 0
    when = ""
    try:
        when = (t["when_postponed"] or "")
    except Exception:
        when = ""

    if postponed == 1 and when:
        try:
            dt = datetime.fromisoformat(when.replace("Z", "+00:00"))
            return f" (был перенос {dt.astimezone(TZINFO).strftime('%d.%m.%Y')})"
        except Exception:
            return " (был перенос)"
    return ""

def send(chat_id, text, markup=None, *, base_delay_s: float = 0.0, _allow_chunk=True):
    """
    Отправка с:
    - уважением rate-limit (sleep base_delay_s),
    - ретраем при 429,
    - fallback-разбиением при 400 'message is too long',
    - логом ошибок и ловлей сетевых исключений.
    """
    if base_delay_s > 0:
        time.sleep(base_delay_s)

    payload = {
        "chat_id": str(chat_id),
        "text": text,
        "disable_web_page_preview": True,
        "parse_mode": "HTML",
    }
    if markup:
        payload["reply_markup"] = markup

    for attempt in range(3):
        try:
            r = requests.post(f"{TG_API}/sendMessage", json=payload, timeout=15)
        except Exception as e:
            print(f"[scheduler.send] EXC chat={chat_id} attempt={attempt+1}: {e}")
            time.sleep(0.5)
            continue

        if r.status_code == 429:
            try:
                retry = r.json().get("parameters", {}).get("retry_after", 1)
            except Exception:
                retry = 1
            time.sleep(retry + 0.5)
            continue

        if r.status_code == 400 and _allow_chunk and "message is too long" in (r.text or "").lower():
            print(f"[scheduler.send] LONG chat={chat_id} -> chunking...")
            send_long(chat_id, text, base_delay_s=0.0)
            return

        if not r.ok:
            print(f"[scheduler.send] FAIL chat={chat_id} code={r.status_code} body={r.text[:300]}")
        else:
            print(f"[scheduler.send] OK chat={chat_id}")
        break

def _chunk_html(text: str, limit: int = 3800) -> list[str]:
    """Режем по двойным \\n\\n (блоки), затем по строкам, и только потом — по символам."""
    chunks, cur = [], ""
    def flush():
        nonlocal cur
        if cur:
            chunks.append(cur)
            cur = ""
    for block in text.split("\n\n"):
        block2 = block + "\n\n"
        if len(block2) <= limit - len(cur):
            cur += block2
            continue
        for line in (block2.split("\n")):
            line2 = line + "\n"
            if len(line2) <= limit - len(cur):
                cur += line2
            else:
                s = line2
                while s:
                    take = min(len(s), limit - len(cur))
                    cur += s[:take]
                    s = s[take:]
                    if len(cur) >= limit:
                        flush()
        if len(cur) >= limit:
            flush()
    flush()
    return [c.strip() for c in chunks if c.strip()]

def send_long(chat_id, text, *, chunk=3800, base_delay_s=0.0):
    parts = _chunk_html(text, limit=chunk)
    total = len(parts)
    for i, part in enumerate(parts, 1):
        head = f"(часть {i}/{total})\n\n" if total > 1 and i > 1 else ""
        send(chat_id, head + part, base_delay_s=base_delay_s, _allow_chunk=False)
        time.sleep(0.3)

def send_document(chat_id, file_path, *, caption=None, base_delay_s: float = 0.0):
    """
    Отправка документа с ретраями 429 и паузами.
    """
    if base_delay_s > 0:
        time.sleep(base_delay_s)

    for attempt in range(3):
        files = {"document": (os.path.basename(file_path), open(file_path, "rb"))}
        data = {"chat_id": str(chat_id)}
        if caption:
            data["caption"] = caption
            data["parse_mode"] = "HTML"
        try:
            r = requests.post(f"{TG_API}/sendDocument", data=data, files=files, timeout=60)
        except Exception as e:
            print(f"[scheduler.sendDocument] EXC chat={chat_id} attempt={attempt+1}: {e}")
            time.sleep(0.5)
            continue
        finally:
            try:
                files["document"][1].close()
            except Exception:
                pass

        if r.status_code == 429:
            try:
                retry = r.json().get("parameters", {}).get("retry_after", 1)
            except Exception:
                retry = 1
            time.sleep(retry + 0.5)
            continue

        if not r.ok:
            print(f"[scheduler.sendDocument] FAIL chat={chat_id} code={r.status_code} body={r.text[:300]}")
        else:
            print(f"[scheduler.sendDocument] OK chat={chat_id} file={os.path.basename(file_path)}")
        break

def _fmt_local(iso_utc: str) -> str:
    if not iso_utc:
        return ""
    try:
        return datetime.fromisoformat(iso_utc.replace("Z", "+00:00")).astimezone(TZINFO).strftime("%d.%m.%Y %H:%M")
    except Exception:
        return iso_utc

def _fmt_local_date(iso_utc: str) -> str:
    if not iso_utc:
        return ""
    try:
        return datetime.fromisoformat(iso_utc.replace("Z", "+00:00")).astimezone(TZINFO).strftime("%d.%m.%Y")
    except Exception:
        return iso_utc

# --------------------------- СТАРЫЙ XLSX/CSV (оставляем как есть) ---------------------------
def build_excel_report_file(now=None) -> str:
    """
    Строит Excel (или CSV при отсутствии openpyxl) с таблицей:
    Человек | ID | Текст задачи | Статус | Когда поставлена | Переносы дедлайна? | История переносов
    Включаем только задачи в статусах open/in_progress.
    Возвращает путь к файлу.
    """
    now = (now or datetime.now(TZINFO))
    stamp = now.strftime("%Y%m%d_%H%M")
    base_name_xlsx = f"daily_tasks_report_{stamp}.xlsx"
    base_name_csv  = f"daily_tasks_report_{stamp}.csv"
    path_xlsx = os.path.join(REPORT_DIR, base_name_xlsx)
    path_csv  = os.path.join(REPORT_DIR, base_name_csv)

    with get_conn() as c:
        rows = c.execute(
            """
            SELECT * FROM tasks
            WHERE status IN ('open','in_progress')
            ORDER BY
            assignee,
            CASE WHEN TRIM(COALESCE(deadline, '')) = '' THEN 1 ELSE 0 END,
            deadline,
            created_at
            """
        ).fetchall()

    header = ["Человек", "ID", "Текст задачи", "Статус", "Когда поставлена", "Переносы дедлайна?", "История переносов"]
    data_rows = []

    for r in rows:
        dchs = get_deadline_changes_for_task(r["id"])
        had_postpone = len(dchs) > 0
        hist = []
        for d in dchs:
            by_who = ""
            try:
                if hasattr(d, "keys") and "by_who" in d.keys():
                    by_who = (d["by_who"] or "").strip()
            except Exception:
                by_who = ""

            when = _fmt_local(d["at"])
            oldd = d["old_deadline"] or "—"
            newd = d["new_deadline"] or "—"
            by = f" (кем: {by_who})" if by_who else ""
            hist.append(f"{oldd} → {newd} [{when}{by}]")

        data_rows.append([
            r["assignee"] or "—",
            r["id"],
            r["task"],
            r["status"],
            _fmt_local(r["created_at"]),
            "Да" if had_postpone else "Нет",
            "; ".join(hist) if hist else "—",
        ])

    if openpyxl:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Задачи в работе"
        for j, val in enumerate(header, 1):
            cell = ws.cell(row=1, column=j, value=val)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for i, row in enumerate(data_rows, 2):
            for j, val in enumerate(row, 1):
                ws.cell(row=i, column=j, value=val)
        widths = [16, 8, 60, 14, 20, 20, 80]
        for idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = w
        wb.save(path_xlsx)
        return path_xlsx

    with open(path_csv, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(header)
        writer.writerows(data_rows)
    return path_csv

# --------------------------- НОВОЕ: персональные PDF-отчёты ---------------------------
# ВНИМАНИЕ: заменить существующую _pdf_draw_wrapped в scheduler.py

from reportlab.pdfbase import pdfmetrics

def _pdf_draw_wrapped(c, text, x_mm, y_mm, *, max_width_mm=170, font_name="Helvetica", font_size=10, line_spacing=1.3):
    """
    Рисует многострочный текст с переносами по реальной ширине.
    - Ширина измеряется pdfmetrics.stringWidth (в pt).
    - Переносим по словам; очень длинные «слова» (URL/без пробелов) рубим жёстко.
    - Возвращает новую координату y (в мм).
    """
    # перевод мм -> pt
    MM_TO_PT = 72.0 / 25.4
    max_width_pt = max_width_mm * MM_TO_PT
    x_pt = x_mm * MM_TO_PT
    y_pt = y_mm * MM_TO_PT

    c.setFont(font_name, font_size)

    def split_hard(word, remain_pt):
        """Рубим очень длинные безпробельные фрагменты так, чтобы куски влазили."""
        out = []
        start = 0
        while start < len(word):
            lo, hi = 1, len(word) - start
            # бинарный поиск максимального вместимого куска
            while lo <= hi:
                mid = (lo + hi) // 2
                chunk = word[start:start+mid]
                w = pdfmetrics.stringWidth(chunk, font_name, font_size)
                if w <= remain_pt:
                    lo = mid + 1
                else:
                    hi = mid - 1
            take = max(1, hi)
            out.append(word[start:start+take])
            start += take
            # на следующий кусок у нас будет полная строка, не остаток
            remain_pt = max_width_pt
        return out

    lines = []
    for para in (text or "").split("\n"):
        words = para.split(" ") if para else [""]
        cur = ""
        cur_w = 0.0

        for w in words:
            # слово + пробел
            token = (w + " ") if w else " "
            token_w = pdfmetrics.stringWidth(token, font_name, font_size)

            if token_w <= (max_width_pt - cur_w):
                cur += token
                cur_w += token_w
            else:
                # если само слово шире строки — рубим жёстко на куски
                if pdfmetrics.stringWidth(w, font_name, font_size) > max_width_pt:
                    # сперва дописываем текущую строку, если в ней что-то есть
                    if cur.strip():
                        lines.append(cur.rstrip())
                        cur, cur_w = "", 0.0
                    chunks = split_hard(w, max_width_pt)
                    for i, ch in enumerate(chunks):
                        if i < len(chunks) - 1:
                            lines.append(ch)
                        else:
                            cur = ch + " "
                            cur_w = pdfmetrics.stringWidth(cur, font_name, font_size)
                else:
                    # перенос по словам
                    if cur.strip():
                        lines.append(cur.rstrip())
                    cur = token
                    cur_w = token_w

        if cur.strip() or para == "":
            lines.append(cur.rstrip())

    # рисуем
    # межстрочный шаг в мм исходя из font_size (pt)
    line_h_mm = (font_size * line_spacing) / (72.0 / 25.4)
    for ln in lines:
        c.drawString(x_pt, y_pt, ln)
        y_pt -= line_h_mm * MM_TO_PT

    return y_pt / MM_TO_PT


def build_combined_pdf_report(now=None) -> str | None:
    """
    Строит ОДИН общий PDF: секции по исполнителям (Имя (@ник)),
    внутри — список всех его open/in_progress задач.
    Возвращает путь к PDF или None, если reportlab недоступен.
    """
    if canvas is None:
        return None

    from db import get_nickname_by_tid, get_reassignments_for_task  # локальный импорт
    now = now or datetime.now(TZINFO)

    # 1) Все задачи в работе
    with get_conn() as c:
        rows = c.execute("""
            SELECT * FROM tasks
            WHERE status IN ('open','in_progress')
            ORDER BY assignee, id
        """).fetchall()

    # 2) Группируем по исполнителю
    by_person: dict[tuple[str, str], list] = {}
    for r in rows:
        key = ((r["assignee"] or "—"), (r["telegram_id"] or ""))
        by_person.setdefault(key, []).append(r)

    # 3) Имя файла
    fname = f"tasks_all_{now.strftime('%Y%m%d_%H%M')}.pdf"
    path = os.path.join(REPORT_DIR, fname)

    # 4) PDF + шрифты
    cpdf = canvas.Canvas(path, pagesize=A4)
    width, height = A4
    x_left = 15  # мм
    y = (height / mm) - 20

    reg, reg_b = _ensure_pdf_font()
    font = reg if reg else "Helvetica"
    font_b = reg_b if reg_b else "Helvetica-Bold"

    cpdf.setTitle("Отчёт по всем исполнителям")
    cpdf.setFont(font_b, 16)
    cpdf.drawString(x_left * mm, y * mm, "Отчёт по всем исполнителям (open / in_progress)")
    y -= 12
    cpdf.setFont(font, 10)

    # 5) По исполнителям
    for (name, tid), tasks in by_person.items():
        # новая страница для шапки секции при нехватке места
        if y < 30:
            cpdf.showPage()
            y = (height / mm) - 20
            cpdf.setFont(font, 10)

        # Шапка секции: Имя (@ник)
        nick = get_nickname_by_tid(tid)
        disp = (name or "—")
        if nick:
            disp += f" ({nick if nick.startswith('@') else '@' + nick})"

        cpdf.setFont(font_b, 12)
        cpdf.drawString(x_left * mm, y * mm, disp)
        y -= 7
        cpdf.setFont(font, 10)

        if not tasks:
            y = _pdf_draw_wrapped(cpdf, "• (задач нет)", x_left, y, max_width_mm=180)
            y -= 4
            continue

        # Список задач
        for t in tasks:
            if y < 25:
                cpdf.showPage()
                y = (height / mm) - 20
                cpdf.setFont(font, 10)

            header = f"• [ID {t['id']}] {t['task']}"
            y = _pdf_draw_wrapped(cpdf, header, x_left, y, max_width_mm=180, font_name=font, font_size=10, line_spacing=1.3)


            dl = (t["deadline"] or "—")
            y = _pdf_draw_wrapped(cpdf, f"Дедлайн: {dl}", x_left + 5, y, max_width_mm=175, font_name=font, font_size=10, line_spacing=1.3)

            dchs = get_deadline_changes_for_task(t["id"])
            y = _pdf_draw_wrapped(cpdf, f"Переносов дедлайна: {len(dchs)}", x_left + 5, y, max_width_mm=175, font_name=font, font_size=10, line_spacing=1.3)


            reas = get_reassignments_for_task(t["id"])
            moves = "; ".join([f"{x['old_assignee']} → {x['new_assignee']}" for x in reas]) if reas else "—"
            y = _pdf_draw_wrapped(cpdf, f"Переназначения: {moves}", x_left + 5, y, max_width_mm=175, font_name=font, font_size=10, line_spacing=1.3)


            y -= 3  # межстрочный отступ между задачами

        y -= 4  # отступ между секциями

    cpdf.save()
    return path


def build_personal_pdf_reports(now=None) -> list[tuple[str, str]]:
    """
    Возвращает список (display_name, pdf_path) для всех исполнителей, у которых есть задачи open/in_progress.
    """
    if canvas is None:
        return []

    from db import get_nickname_by_tid, get_reassignments_for_task
    now = now or datetime.now(TZINFO)

    with get_conn() as c:
        rows = c.execute("""
            SELECT * FROM tasks
            WHERE status IN ('open','in_progress')
            ORDER BY assignee, id
        """).fetchall()

    by_person = {}
    for r in rows:
        key = (r["assignee"] or "—", r["telegram_id"] or "")
        by_person.setdefault(key, []).append(r)

    out = []
    for (name, tid), tasks in by_person.items():
        if not tasks:
            continue

        nick = get_nickname_by_tid(tid)
        disp = name or "—"
        if nick:
            disp += f" ({nick if nick.startswith('@') else '@'+nick})"

        filename = f"tasks_{(name or 'no_name').replace(' ', '_')}_{now.strftime('%Y%m%d_%H%M')}.pdf"
        path = os.path.join(REPORT_DIR, filename)

        cpdf = canvas.Canvas(path, pagesize=A4)
        reg, reg_b = _ensure_pdf_font()
        font = reg or "Helvetica"
        font_b = reg_b or "Helvetica-Bold"

        width, height = A4
        x_left = 15
        y = (height / mm) - 20

        cpdf.setFont(font_b, 16)
        cpdf.drawString(x_left * mm, y * mm, disp)
        y -= 10
        cpdf.setFont(font, 10)

        for t in tasks:
            if y < 20:
                cpdf.showPage()
                y = (height / mm) - 20
                cpdf.setFont(font, 10)

            header = f"• [ID {t['id']}] {t['task']}"
            y = _pdf_draw_wrapped(cpdf, header, x_left, y, max_width_mm=180, font_name=font, font_size=10, line_spacing=1.3)


            dl = (t["deadline"] or "—")
            y = _pdf_draw_wrapped(cpdf, f"Дедлайн: {dl}", x_left + 5, y, max_width_mm=175, font_name=font, font_size=10, line_spacing=1.3)


            dchs = get_deadline_changes_for_task(t["id"])
            y = _pdf_draw_wrapped(cpdf, f"Переносов дедлайна: {len(dchs)}", x_left + 5, y, max_width_mm=175, font_name=font, font_size=10, line_spacing=1.3)


            reas = get_reassignments_for_task(t["id"])
            moves = "; ".join([f"{x['old_assignee']} → {x['new_assignee']}" for x in reas]) if reas else "—"
            y = _pdf_draw_wrapped(cpdf, f"Переназначения: {moves}", x_left + 5, y, max_width_mm=175, font_name=font, font_size=10, line_spacing=1.3)


            y -= 3

        cpdf.save()
        out.append((disp, path))

    return out

# --------------------------- Текстовые отчёты/напоминалки ---------------------------
def build_admin_text(now=None) -> str:
    now = (now or datetime.now(TZINFO)).replace(minute=0, second=0, microsecond=0)
    if now.weekday() not in WEEKDAYS:
        return "Выходной день. Отчёт не формируется."

    day_start_local = now.replace(hour=0)
    day_end_local   = now.replace(hour=23, minute=59, second=59)
    day_start_utc = day_start_local.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    day_end_utc   = day_end_local.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    total_open_like = count_open_like()
    closed_today = count_closed_between(day_start_utc, day_end_utc)

    overdue = get_overdue_open_tasks(now.strftime("%Y-%m-%d"))
    overdue_count = len(overdue)

    lines_overdue = []
    for t in overdue:
        dl = _fmt_local_date(t["deadline"]) if (t["deadline"] or "").strip() else "—"
        lines_overdue.append(f"{t['id']}; {t['task']}; {t['assignee'] or '—'}; {dl}")

    admin = []
    admin.append("🧾 Ежедневный отчёт (конец дня)")
    admin.append("")
    admin.append(f"Задач в работе всего: {total_open_like} (open + in_progress)")
    admin.append(f"Выполнено за сегодня: {closed_today}")
    admin.append(f"Просрочено задач: {overdue_count}")
    admin.append("")
    admin.append("<b>Список просроченных задач:</b>")
    admin.append("ID; Описание; Исполнитель; Дедлайн")
    admin.append(("(нет)" if not lines_overdue else "\n".join(lines_overdue)))
    admin.append("")
    admin.append("Подробный отчёт см. в прикреплённом файле.")
    return "\n".join(admin)

def kb_reminder(task_id: int):
    return {
        "inline_keyboard": [[
            {"text": "✅ Я сделал", "callback_data": f"done:{task_id}"},
            {"text": "⏰ Я не успеваю", "callback_data": f"cant_do:{task_id}"}
        ]]
    }

def jobs_tick():
    """Запускать раз в час."""
    now = datetime.now(TZINFO).replace(minute=0, second=0, microsecond=0)

    # 0) Если рабочее окно — выгружаем дозревшее из outbox
    if is_work_time(now):
        due = pop_due_outbox(datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"))
        for row in due:
            try:
                markup = json.loads(row["markup"]) if row["markup"] else None
            except Exception:
                markup = None
            try:
                send(row["chat_id"], row["text"], markup)
                mark_outbox_sent(row["id"])
            except Exception:
                pass

    # 1) Вытащим все НЕ закрытые задачи
    with get_conn() as c:
        rows = c.execute(
            "SELECT * FROM tasks WHERE status IN ('open','in_progress')"
        ).fetchall()

    # 1a) Просрочки — будни в 10:00
    if now.weekday() in WEEKDAYS and now.hour == 10:
        for r in rows:
            dl = (r["deadline"] or "").strip()
            if not dl:
                continue
            try:
                d = datetime.strptime(dl, "%Y-%m-%d").date()
                if d < now.date():
                    text = (
                        "⛔️ Просроченная задача.\n\n"
                        f"🧩 <b>{h(r['task'])}</b>\n"
                        f"📅 Дедлайн был: {dl}\n\n"
                        f"ID: #{r['id']}\n"
                        "Когда сможешь закрыть? Ответь датой или нажми «⏰ Я не успеваю»."
                    )
                    send_or_enqueue(r["telegram_id"], text, kb_reminder(r["id"]))
            except Exception:
                pass

    # 2) Напоминалки после initial (3-й день), «за день» и «в день» дедлайна
    for r in rows:
        chat = r["telegram_id"]
        task_id = r["id"]
        task = r["task"]
        deadline = r["deadline"] or ""
        initial = r["initial_text_sent"]
        if not initial:
            continue
        try:
            initial_dt = datetime.fromisoformat(initial.replace("Z", "+00:00")).astimezone(TZINFO)
        except Exception:
            continue

        three_days = initial_dt + timedelta(days=3)
        if three_days <= now < three_days + timedelta(hours=1):
            text = (
                "Эй, помнишь про эту задачу?\n\n"
                f"🧩 <b>{h(task)}</b>"
                f"📅 Дедлайн: {deadline or '—'}\n\n"
                "Когда планируешь добить? Ответь датой (например, 2025-09-01) или «завтра/послезавтра»."
            )
            try:
                send_or_enqueue(chat, text, kb_reminder(task_id))
            except Exception:
                pass

        if deadline:
            try:
                d = datetime.strptime(deadline, "%Y-%m-%d").replace(tzinfo=TZINFO)
                if now.date() == (d - timedelta(days=1)).date() and now.hour == 10:
                    text = (
                        "Напоминаю: завтра дедлайн по задаче:\n\n"
                        f"🧩 <b>{task}</b>\n\n"
                        "Успеваешь? Если нет — ткни кнопку надо команду поставить, перенесём цивилизованно."
                    )
                    send_or_enqueue(chat, text, kb_reminder(task_id))
            except Exception:
                pass

        if deadline:
            try:
                d = datetime.strptime(deadline, "%Y-%m-%d").replace(tzinfo=TZINFO)
                if now.date() == d.date() and now.hour == 10:
                    text = (
                        "Сегодня дедлайн. Как там дела?\n\n"
                        f"🧩 <b>{task}</b>"
                    )
                    send_or_enqueue(chat, text, kb_reminder(task_id))
            except Exception:
                pass

    # 3) Ежедневная сводка (будни 18:00)
    if (now.hour == 18) and (now.weekday() in WEEKDAYS):
        assignees = list_unique_assignees()
        sent_count = 0
        for name, chat in assignees:
            if not str(chat).strip():
                continue
            tasks_open = find_open_tasks_for_user(chat)
            if not tasks_open:
                continue
            lines = []
            for i, t in enumerate(tasks_open, 1):
                mark = format_postponed(t)
                tail = f" — до {t['deadline']}" if t['deadline'] else ""
                lines.append(f"{i}. {t['task']}{tail}{mark}")
            user_msg = "📋 Задачи в работе на конец дня:\n" + "\n".join(lines)
            try:
                send_or_enqueue(chat, user_msg, base_delay_s=ASSIGNEE_DELAY_S)
                sent_count += 1
                if sent_count % ASSIGNEE_BATCH_SIZE == 0:
                    time.sleep(ASSIGNEE_BATCH_PAUSE_S)
            except Exception as e:
                print(f"[scheduler.user_digest] ERROR send to {chat}: {e}")

def build_admin_text_only(now=None) -> str:
    return build_admin_text(now)

def send_admin_report_now(now=None):
    """
    Ручной прогон: шлём короткий текст + прикрепляем Excel/CSV всем адресатам.
    """
    now = now or datetime.now(TZINFO)
    admin_text = build_admin_text(now)
    report_path = build_excel_report_file(now)

    recipients = [str(VADIM_CHAT_ID), *map(str, ASSISTANT_CHAT_IDS)]
    seen = set()
    for idx, cid in enumerate(recipients, 1):
        if cid in seen or not cid.strip():
            continue
        try:
            # вне рабочего окна НЕ откладываем — ручной форс
            send(cid, admin_text, base_delay_s=(0.0 if idx == 1 else ASSISTANT_DELAY_S))
        except Exception as e:
            print(f"[scheduler.force] ERROR text to {cid}: {e}")
        try:
            send_document(cid, report_path, caption="📎 Прикреплён файл отчёта.", base_delay_s=0.2)
        except Exception as e:
            print(f"[scheduler.force] ERROR doc to {cid}: {e}")
        seen.add(cid)

def main():
    # простой «ежечасный» цикл
    while True:
        try:
            jobs_tick()
        except Exception as e:
            print("scheduler error:", e)
        # спим до следующего часа
        now = datetime.now(TZINFO)
        sleep_secs = 3600 - (now.minute * 60 + now.second)
        time.sleep(max(30, sleep_secs))

if __name__ == "__main__":
    main()
